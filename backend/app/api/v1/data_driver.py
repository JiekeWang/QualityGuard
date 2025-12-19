"""
数据驱动文件导入API
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import Optional, List, Any, Dict
import csv
import io
import json

from app.core.database import get_db
from app.core.dependencies import get_current_active_user
from app.models.user import User
from app.models.test_case import TestCase

router = APIRouter()


def _parse_csv_content(content: str) -> tuple[List[Dict[str, Any]], List[str]]:
    """解析CSV内容
    
    Args:
        content: CSV文件内容
    
    Returns:
        (数据列表, 错误列表)
    """
    data_list = []
    errors = []
    
    try:
        # 使用csv.DictReader解析CSV
        csv_file = io.StringIO(content)
        reader = csv.DictReader(csv_file)
        
        if not reader.fieldnames:
            errors.append("CSV文件为空或格式不正确")
            return data_list, errors
        
        for row_index, row in enumerate(reader, start=2):  # 从第2行开始（第1行是标题）
            # 过滤空行
            if not any(row.values()):
                continue
            
            # 清理空值
            row_data = {k: v.strip() if isinstance(v, str) else v for k, v in row.items() if k and v}
            
            if row_data:
                row_data['__row_index'] = row_index
                data_list.append(row_data)
        
        if not data_list:
            errors.append("CSV文件中没有有效数据")
    
    except Exception as e:
        errors.append(f"CSV解析失败: {str(e)}")
    
    return data_list, errors


def _parse_excel_content(file_content: bytes, sheet_name: Optional[str] = None) -> tuple[List[Dict[str, Any]], List[str]]:
    """解析Excel内容
    
    Args:
        file_content: Excel文件内容
        sheet_name: 工作表名称（可选）
    
    Returns:
        (数据列表, 错误列表)
    """
    data_list = []
    errors = []
    
    try:
        import openpyxl
        from io import BytesIO
        
        # 加载Excel文件
        workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        
        # 选择工作表
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        
        if not worksheet:
            errors.append("Excel文件中没有有效的工作表")
            return data_list, errors
        
        # 获取标题行（第一行）
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            errors.append("Excel文件为空")
            return data_list, errors
        
        headers = rows[0]
        if not headers or not any(headers):
            errors.append("Excel文件缺少标题行")
            return data_list, errors
        
        # 解析数据行
        for row_index, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue  # 跳过空行
            
            # 构建数据字典
            row_data = {}
            for col_index, (header, value) in enumerate(zip(headers, row)):
                if header and value is not None:
                    # 转换为字符串
                    if isinstance(value, (int, float)):
                        row_data[str(header)] = value
                    else:
                        row_data[str(header)] = str(value).strip() if value else ''
            
            if row_data:
                row_data['__row_index'] = row_index
                data_list.append(row_data)
        
        if not data_list:
            errors.append("Excel文件中没有有效数据")
        
        workbook.close()
    
    except ImportError:
        errors.append("服务器未安装openpyxl库，无法解析Excel文件")
    except Exception as e:
        errors.append(f"Excel解析失败: {str(e)}")
    
    return data_list, errors


def _validate_data(data_list: List[Dict[str, Any]]) -> tuple[int, int, List[str]]:
    """验证数据
    
    Args:
        data_list: 数据列表
    
    Returns:
        (有效行数, 无效行数, 警告列表)
    """
    valid_count = 0
    invalid_count = 0
    warnings = []
    
    for data in data_list:
        # 基本验证：至少包含一个非 expected_* 和非 __row_index 的字段
        has_data_field = any(k for k in data.keys() if not k.startswith('expected_') and not k.startswith('__'))
        
        if has_data_field:
            valid_count += 1
        else:
            invalid_count += 1
            row_index = data.get('__row_index', '未知')
            warnings.append(f"第{row_index}行：没有有效的数据字段")
    
    return valid_count, invalid_count, warnings


@router.post("/test-cases/{case_id}/data-driver/import")
async def import_data_driver_file(
    case_id: int,
    file: UploadFile = File(...),
    preview: bool = Form(False),
    sheet_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导入CSV/Excel文件作为数据驱动测试数据
    
    Args:
        case_id: 测试用例ID
        file: 上传的文件（CSV或Excel）
        preview: 是否仅预览（不保存）
        sheet_name: Excel工作表名称（可选）
    """
    # 检查测试用例是否存在
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    test_case = result.scalar_one_or_none()
    if not test_case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 读取文件内容
    try:
        file_content = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件读取失败: {str(e)}")
    
    # 根据文件类型解析
    filename = file.filename or ""
    data_list = []
    errors = []
    
    if filename.endswith('.csv'):
        # 解析CSV
        try:
            content = file_content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                content = file_content.decode('gbk')
            except:
                raise HTTPException(status_code=400, detail="CSV文件编码不支持，请使用UTF-8或GBK编码")
        
        data_list, errors = _parse_csv_content(content)
    
    elif filename.endswith(('.xlsx', '.xls')):
        # 解析Excel
        data_list, errors = _parse_excel_content(file_content, sheet_name)
    
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式，仅支持CSV和Excel文件")
    
    if errors:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {'; '.join(errors)}")
    
    # 验证数据
    valid_count, invalid_count, warnings = _validate_data(data_list)
    
    # 如果是预览模式，返回前100条数据
    if preview:
        preview_data = data_list[:100] if len(data_list) > 100 else data_list
        return {
            "success": True,
            "preview": True,
            "data": preview_data,
            "summary": {
                "total_rows": len(data_list),
                "valid_rows": valid_count,
                "invalid_rows": invalid_count,
                "warnings": warnings,
                "preview_count": len(preview_data)
            }
        }
    
    # 非预览模式：保存数据
    # 移除 __row_index 字段
    clean_data = []
    for data in data_list:
        clean_item = {k: v for k, v in data.items() if not k.startswith('__')}
        clean_data.append(clean_item)
    
    # 更新测试用例的data_driver配置
    try:
        # 获取当前配置
        current_config = test_case.config or {}
        if isinstance(current_config, str):
            current_config = json.loads(current_config)
        
        # 获取当前data_driver配置
        data_driver = current_config.get('data_driver', {})
        if isinstance(data_driver, str):
            data_driver = json.loads(data_driver)
        
        # 更新data字段
        data_driver['data'] = clean_data
        data_driver['metadata'] = {
            "total_rows": len(clean_data),
            "imported_at": datetime.now().isoformat(),
            "format": "excel" if filename.endswith(('.xlsx', '.xls')) else "csv",
            "filename": filename
        }
        
        # 保存回config
        current_config['data_driver'] = data_driver
        test_case.config = current_config
        
        # 标记为数据驱动
        test_case.is_data_driven = True
        
        await db.commit()
        await db.refresh(test_case)
        
        return {
            "success": True,
            "preview": False,
            "saved_count": len(clean_data),
            "summary": {
                "total_rows": len(clean_data),
                "valid_rows": valid_count,
                "invalid_rows": invalid_count,
                "warnings": warnings
            },
            "data_driver": data_driver
        }
    
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"保存失败: {str(e)}")


@router.post("/test-cases/{case_id}/data-driver/save")
async def save_data_driver(
    case_id: int,
    data: Dict[str, Any],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """保存数据驱动测试数据
    
    Args:
        case_id: 测试用例ID
        data: 数据驱动配置
            {
                "data": [...],  # 数据列表
                "replace": true  # 是否替换现有数据
            }
    """
    # 检查测试用例是否存在
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    test_case = result.scalar_one_or_none()
    if not test_case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    try:
        # 获取当前配置
        current_config = test_case.config or {}
        if isinstance(current_config, str):
            current_config = json.loads(current_config)
        
        # 获取当前data_driver配置
        data_driver = current_config.get('data_driver', {})
        if isinstance(data_driver, str):
            data_driver = json.loads(data_driver)
        
        # 获取新数据
        new_data = data.get('data', [])
        replace = data.get('replace', True)
        
        # 更新或追加数据
        if replace:
            data_driver['data'] = new_data
        else:
            existing_data = data_driver.get('data', [])
            data_driver['data'] = existing_data + new_data
        
        # 更新元数据
        from datetime import datetime
        data_driver['metadata'] = {
            "total_rows": len(data_driver['data']),
            "updated_at": datetime.now().isoformat(),
            "format": "manual"
        }
        
        # 保存回config
        current_config['data_driver'] = data_driver
        test_case.config = current_config
        
        # 标记为数据驱动
        test_case.is_data_driven = True
        
        await db.commit()
        await db.refresh(test_case)
        
        return {
            "success": True,
            "saved_count": len(data_driver['data']),
            "data_driver": data_driver
        }
    
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"保存失败: {str(e)}")

